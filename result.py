import json
from typing import Dict


def get_results(json_filename: str) -> Dict[str, str]:
    """Compute winners per post from a data file."""
    with open(json_filename, "r", encoding="utf-8") as json_file:
        data = json.load(json_file)

    result_dict: Dict[str, str] = {}
    for entry in data:
        post = entry["post"]
        candidates = entry["candidates"]
        if not candidates:
            continue
        max_votes_candidate = max(candidates, key=lambda x: x.get("votes", 0))
        result_dict[post] = max_votes_candidate.get("name", "")

    return result_dict


from openpyxl import Workbook


def create_excel(results: Dict[str, str], excel_filename: str = "results.xlsx") -> str:
    """Create an Excel file from results and return the filename."""
    workbook = Workbook()
    sheet = workbook.active

    sheet["A1"] = "Post"
    sheet["B1"] = "Winner"

    row = 2
    for post, candidate in results.items():
        sheet.cell(row=row, column=1, value=post)
        sheet.cell(row=row, column=2, value=candidate)
        row += 1

    workbook.save(excel_filename)
    return excel_filename


if __name__ == "__main__":
    # Simple CLI usage without progress bars
    json_filename = "templates/Data/Data.json"
    results = get_results(json_filename)
    outfile = create_excel(results)
    print(f"Results saved to {outfile}")
